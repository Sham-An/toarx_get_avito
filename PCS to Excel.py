import pandas as pd
import openpyxl
import subprocess
from openpyxl.styles import Border, Side
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles import NamedStyle, Font, Border, Side

#>>> subprocess.Popen('C:\\Windows\\System32\\calc.exe')
# Метод poll() возвращает значение None, если в момент его вызова процесс все еще выполняется. Если же процесс к этому моменту завершен, то он возвращает код завершения процесса. Код заверешения служит индикатором того, завершился ли процесс без ошибок (код равен нулю) или же его завершение было вызвано ошибкой (ненулевой код).
# Метод wait() ожидает завершения запущенного процесса, прежде чем продолжить выполнение основной программы. Возвращаемым значением метода является целочисленный код завершения процесса.
# >>> notepad = subprocess.Popen('C:\\Windows\\System32\\notepad.exe')
# >>> notepad.poll() == None
# True
# >>> notepad.poll() == None
# False
# >>> notepad.wait()
# 0
# >>> notepad.poll()
# 0
# Двойной клик на иконке файла с расширением .txt позволяет автоматически запустить приложение, ассоциированное с этим расширением. Функция Popen() также может открывать файлы подобным образом:
#
# >>> subprocess.Popen(('start', 'C:\\example\\readme.txt'), shell = True)
# subprocess.Popen(('start', 'pandas_to_excel.xlsx'), shell = True)
# <subprocess.Popen object at 0x0000020B183EDE10>Копировать
# В каждой операционной системе есть программа, выполняющая те же функции, что и двойной клик на иконке файла. В Windows это программа start, в Ubuntu Linux — программа see.
#
# Именованный аргумент shell = True нужен только для ОС Windows.



dict_list = [
            {'PCS_id': 3982, 'kod_start': '20191', 'Log_Time': '2020-12-29 13:11:04.663', 'start': '2020-12-29 13:11:04.897', 'kod_stop': '29321', 'stop': '2020-12-29 13:47:40.445', 'parent_group_file': 'ФЛУКОНАЗОЛ', 'kalibr': '500', 'kog_group': '000000039', 'file_dir': 'Флуконазол 150', 'Full_path': 'FULL PATH', 'file_name': 'Флуконазол 500-000000039.VDF', 'GTIN_kod': '--', 'GTIN_name': '--', 'last_kod_id': '--', 'last_kod': '--', 'status': '--'},
            {'PCS_id': 3991, 'kod_start': '29322', 'Log_Time': '2020-12-29 13:52:19.974', 'start': '2020-12-29 13:52:20.192', 'kod_stop': '30499', 'stop': '2020-12-29 13:57:16.935', 'parent_group_file': 'ФЛУКОНАЗОЛ', 'kalibr': '500', 'kog_group': '000000039', 'file_dir': 'Флуконазол 150', 'Full_path': 'FULL PATH', 'file_name': 'Флуконазол 500-000000039.VDF', 'GTIN_kod': '--', 'GTIN_name': '--', 'last_kod_id': '--', 'last_kod': '--', 'status': '--'},
            {'PCS_id': 3998, 'kod_start': '30500', 'Log_Time': '2020-12-29 15:19:49.869', 'start': '2020-12-29 15:19:50.073', 'kod_stop': '31506', 'stop': '2020-12-29 15:24:15.198', 'parent_group_file': 'ФЛУКОНАЗОЛ', 'kalibr': '500', 'kog_group': '000000039', 'file_dir': 'Флуконазол 150', 'Full_path': 'FULL PATH', 'file_name': 'Флуконазол 500-000000039.VDF', 'GTIN_kod': '--', 'GTIN_name': '--', 'last_kod_id': '--', 'last_kod': '--', 'status': '--'},
            ]

df = pd.DataFrame([[11, 21, 31], [12, 22, 32], [31, 32, 33]],
                  index=['one', 'two', 'three'], columns=['a', 'b', 'c'])



#
    # {‘Student_Name’: ‘Samreena’, ‘Course_Title’: ‘SQA’, ‘GPA’: 3.1},
    # {‘Student_Name’: ‘Raees’, ‘Course_Title’: ‘SRE’, ‘GPA’: 3.3},
    # {‘Student_Name’: ‘Sara’, ‘Course_Title’: ‘IT Basics’, ‘GPA’: 2.8},
    # {‘Student_Name’: ‘Sana’, ‘Course_Title’: ‘Artificial Intelligence’, ‘GPA’: 4.0}
    # ]

# Create the DataFrame


def prnDataframe():
    df = pd.DataFrame([[11, 21, 31], [12, 22, 32], [31, 32, 33]],
                   index=['one', 'two', 'three'], columns=['a', 'b', 'c'])
    #
    # print(df)
    # Create a list of dictionaries

    # Create the DataFrame
    dframe = pd.DataFrame(dict_list)
    print(dframe)

def editCells():
    wb = openpyxl.load_workbook('pandas_to_excel.xlsx')
    # Call a Workbook() function of openpyxl
    # to create a new blank Workbook object
    #wb = openpyxl.Workbook()

    # Get workbook active sheet
    # from the active attribute.
    for i, s in enumerate(wb.sheetnames):
        print(s)
        # if s == 'charlie':
        #     break
    #wb.active = i


    wb.active = wb['sheet_dframe']
    sheet = wb.active

    # writing to the specified cell
    #sheet.cell(row=1, column=1).value = ' hello '
    sheet.cell(row=1, column=1).value = '№'
    sheet.cell(row=1, column=2).value = 'PCS'
    sheet.cell(row=1, column=3).value = 'start'
    sheet.cell(row=1, column=6).value = 'stop'
    sheet.cell(row=1, column=7).value = 'время'


    # определим стили сторон
    thins = Side(border_style="medium", color="0000ff")
    double = Side(border_style="dashDot", color="ff0000")
    # рисуем границы
    ws = wb.active
    cell = ws['B2']
    cell.border = Border(top=double, bottom=double, left=thins, right=thins)
    #wb.save("styled_border.xlsx")

    # set the height of the row
    #sheet.row_dimensions[1].height = 70 #высота строки
    # создание переменной именованного стиля
    name_style = NamedStyle(name="highlight5")
    # применение стилей к созданной переменной
    name_style.font = Font(bold=True, size=20)
    bd = Side(style='thick', color="000000")
    name_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    #После создания именованного стиля его нужно зарегистрировать в рабочей книге:
    #if 'highlight4' not in ws.style_names:
    #    ws.add_named_style(name_style)

    # Directly Use as:
    #ws['D5'].style = 'NormalBorderStyle'
    #wb.add_named_style(name_style)
    #Именованные стили также будут автоматически зарегистрированы при первом назначении их ячейке:

    ws['A1'].style = name_style
    #После регистрации стиля в рабочей книге, применять его можно только по имени:
    ws['D5'].style = 'highlight5'


    # set the width of the column
    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 5
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 1
    sheet.column_dimensions['E'].width = 1
    sheet.column_dimensions['F'].width = 10
    sheet.column_dimensions['G'].width = 16
    sheet.column_dimensions['J'].width = 20
    sheet.column_dimensions['H'].width = 20


    # save the file
    # выравниваем текст в ячейках стилями
    # >> > ws['A1'].alignment = Alignment(horizontal='left')
    # >> > ws['A2'].alignment = Alignment(horizontal='center')
    # >> > ws['A3'].alignment = Alignment(horizontal='right')
    #>> > ws = wb.active
    # объединим ячейки в диапазоне `B2:E2`
    #>> > ws.merge_cells('B2:E2')
    # в данном случае крайняя верхняя-левая ячейка это `B2`
    #>> > megre_cell = ws['B2']
    # запишем в нее текст
    #>> > megre_cell.value = 'Объединенные ячейки `B2 : E2`'
    # установить высоту строки
    #>> > ws.row_dimensions[2].height = 30
    # установить ширину столбца
    #>> > ws.column_dimensions['B'].width = 40
    # выравнивание текста
    #>> > megre_cell.alignment = Alignment(horizontal="center", vertical="center")
    # определим стили сторон
    #>> > thins = Side(border_style="medium", color="0000ff")
    #>> > double = Side(border_style="dashDot", color="ff0000")
    # рисуем границы
    #>> > cell.border = Border(top=double, bottom=double, left=thins, right=thins)


    wb.save('pandas_to_excel.xlsx')
    subprocess.Popen(('start', 'pandas_to_excel.xlsx'), shell=True)



def pandasToExcel():
    df2 = df[['a', 'c']]
    dframe = pd.DataFrame(dict_list)
    #print(df2)
    with pd.ExcelWriter('pandas_to_excel.xlsx') as writer:
        df.to_excel(writer, sheet_name='sheet1')
        df2.to_excel(writer, sheet_name='sheet2')
        dframe.to_excel(writer, sheet_name='sheet_dframe')

    subprocess.Popen(('start', 'pandas_to_excel.xlsx'), shell=True)

def Styles():
    wb = Workbook()
    ws = wb.active

    # определим стили сторон
    thins = Side(border_style="thin", color="0000ff")
    double = Side(border_style="double", color="ff0000")

    # начинаем заполнение области ячеек 10x10 данными
    # при этом будем стилизировать границы области
    for r, row in enumerate(range(5, 15), start=1):
        for c, col in enumerate(range(5, 15), start=1):
            # это значение, которое будем записывать в ячейку
            val_cell = r * c
            # левая верхняя ячейка
            if r == 1 and c == 1:
                ws.cell(row=row, column=col, value=val_cell).border = Border(top=double, left=thins)
            # правая верхняя ячейка
            elif r == 1 and c == 10:
                ws.cell(row=row, column=col, value=val_cell).border = Border(top=double, right=thins)
            # верхние ячейки
            if r == 1 and c != 1 and c != 10:
                ws.cell(row=row, column=col, value=val_cell).border = Border(top=double)
            # левая нижняя ячейка
            elif r == 10 and c == 1:
                ws.cell(row=row, column=col, value=val_cell).border = Border(bottom=double, left=thins)
            # правая нижняя ячейка
            elif r == 10 and c == 10:
                ws.cell(row=row, column=col, value=val_cell).border = Border(bottom=double, right=thins)
            # нижние ячейки
            elif r == 10 and c != 1 and c != 10:
                ws.cell(row=row, column=col, value=val_cell).border = Border(bottom=double)
            # левые ячейки
            elif c == 1 and r != 1 and r != 10:
                ws.cell(row=row, column=col, value=val_cell).border = Border(left=thins)
            # правые ячейки
            elif c == 10 and r != 1 and r != 10:
                ws.cell(row=row, column=col, value=val_cell).border = Border(right=thins)
            else:
                # здесь ячейки просто заполняются данными
                ws.cell(row=row, column=col, value=val_cell)

    # сохраняем и смотрим что получилось
    wb.save("styled_border.xlsx")

def Stylecolor():

    #from openpyxl import Workbook
    #from openpyxl.styles import PatternFill, Font, Alignment
    wb = Workbook()
    ws = wb.active

    # объединим ячейки в диапазоне `B2:E2`
    ws.merge_cells('B2:E2')
    megre_cell = ws['B2']
    # запишем в нее текст
    megre_cell.value = 'Объединенные ячейки `B2 : E2`'
    # установить высоту строки
    ws.row_dimensions[2].height = 30
    # установить ширину столбца
    ws.column_dimensions['B'].width = 40
    # заливка ячейки цветом
    megre_cell.fill = PatternFill('solid', fgColor="DDDDDD")
    # шрифт и цвет текста ячейки
    megre_cell.font = Font(bold=True, color='FF0000', name='Arial', size=14)
    # ну и для красоты выровним текст
    megre_cell.alignment = Alignment(horizontal='center', vertical='center')

    #Создание именованного стиля.
    #from openpyxl.styles import NamedStyle, Font, Border, Side

    #wb = Workbook()
    #ws = wb.active

    # создание переменной именованного стиля
    name_style = NamedStyle(name="highlight3")
    # применение стилей к созданной переменной
    name_style.font = Font(bold=True, size=20)
    bd = Side(style='thick', color="000000")
    name_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    #После создания именованного стиля его нужно зарегистрировать в рабочей книге:

    wb.add_named_style(name_style)
    #Именованные стили также будут автоматически зарегистрированы при первом назначении их ячейке:

    ws['A10'].style = name_style
    #После регистрации стиля в рабочей книге, применять его можно только по имени:
    ws['D5'].style = 'highlight3'

    # сохраняем и смотрим что получилось

    wb.save("cell_color.xlsx")

def namedStyle():
    #Создание именованного стиля.
    #from openpyxl.styles import NamedStyle, Font, Border, Side
    wb = Workbook()
    ws = wb.active

    # создание переменной именованного стиля
    name_style = NamedStyle(name="highlight2")
    # применение стилей к созданной переменной
    name_style.font = Font(bold=True, size=20)
    bd = Side(style='thick', color="000000")
    name_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    #После создания именованного стиля его нужно зарегистрировать в рабочей книге:

    wb.add_named_style(name_style)
    #Именованные стили также будут автоматически зарегистрированы при первом назначении их ячейке:

    ws['A1'].style = name_style
    #После регистрации стиля в рабочей книге, применять его можно только по имени:
    ws['D5'].style = 'highlight2'


if __name__ == '__main__':

    #prnDataframe()
    #pandasToExcel()
    editCells()
    #Styles()
    #Stylecolor()
